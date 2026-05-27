"""
STEP 4: GENERATE EXCEL REPORT
==============================
What this does:
- Reads grading results from Step 3
- Creates a formatted Excel file with all marks and deductions
- Saved to grading_output/grades.xlsx

Run: python scripts/04_generate_report.py
"""

import sys
import json
from pathlib import Path
from typing import List, Dict

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.config_loader import config, get_path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_report(results: List[dict], output_path: Path):
    """
    Create a well-formatted Excel report from grading results.
    """
    wb = openpyxl.Workbook()
    
    # ---------- Sheet 1: Summary ----------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title row
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "GRADING SUMMARY REPORT"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Student ID", "Total Marks", "Questions Graded", "Deduction Details"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data rows
    for row_idx, result in enumerate(results, 4):
        num_questions = len(result.get("per_question", {}))
        
        ws_summary.cell(row=row_idx, column=1, value=result["student_id"])
        ws_summary.cell(row=row_idx, column=2, value=result["total_marks"])
        ws_summary.cell(row=row_idx, column=3, value=num_questions)
        ws_summary.cell(row=row_idx, column=4, value=result["deduction_summary"])
        
        # Wrap text for deduction column
        ws_summary.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 80
    
    # ---------- Sheet 2: Detailed Breakdown ----------
    ws_detail = wb.create_sheet("Detailed Breakdown")
    
    # Find all unique question numbers
    all_q_nums = set()
    for result in results:
        all_q_nums.update(result.get("per_question", {}).keys())
    all_q_nums = sorted(all_q_nums)
    
    # Headers
    detail_headers = ["Student ID"]
    for q_num in all_q_nums:
        detail_headers.append(f"Q{q_num} Marks")
        detail_headers.append(f"Q{q_num} Deductions")
    detail_headers.append("Total")
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data
    for row_idx, result in enumerate(results, 2):
        ws_detail.cell(row=row_idx, column=1, value=result["student_id"])
        
        col = 2
        for q_num in all_q_nums:
            q_info = result.get("per_question", {}).get(q_num, {})
            marks = q_info.get("marks", "N/A")
            deductions = "; ".join(q_info.get("deductions", ["N/A"]))
            
            ws_detail.cell(row=row_idx, column=col, value=marks)
            ws_detail.cell(row=row_idx, column=col+1, value=deductions)
            ws_detail.cell(row=row_idx, column=col+1).alignment = Alignment(wrap_text=True)
            col += 2
        
        ws_detail.cell(row=row_idx, column=col, value=result["total_marks"])
    
    # Adjust column widths for detail sheet
    ws_detail.column_dimensions['A'].width = 20
    for i in range(2, len(detail_headers) + 1):
        if i % 2 == 0:  # Marks columns
            ws_detail.column_dimensions[get_column_letter(i)].width = 12
        else:  # Deduction columns
            ws_detail.column_dimensions[get_column_letter(i)].width = 40
    
    # ---------- Sheet 3: Statistics ----------
    ws_stats = wb.create_sheet("Statistics")
    
    # Calculate stats
    if results:
        total_students = len(results)
        total_marks_list = [r["total_marks"] for r in results]
        avg_marks = sum(total_marks_list) / total_students
        max_marks_obtained = max(total_marks_list)
        min_marks_obtained = min(total_marks_list)
        
        stats_data = [
            ["Statistic", "Value"],
            ["Total Students", total_students],
            ["Average Marks", round(avg_marks, 2)],
            ["Highest Marks", max_marks_obtained],
            ["Lowest Marks", min_marks_obtained],
            ["", ""],
            ["Question-wise Averages", ""],
        ]
        
        # Per-question averages
        for q_num in all_q_nums:
            q_marks = []
            for r in results:
                q_info = r.get("per_question", {}).get(q_num, {})
                if "marks" in q_info:
                    q_marks.append(q_info["marks"])
            if q_marks:
                q_avg = sum(q_marks) / len(q_marks)
                stats_data.append([f"Question {q_num}", round(q_avg, 2)])
        
        for row_idx, (label, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row_idx, column=1, value=label)
            ws_stats.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws_stats.cell(row=1, column=1).font = Font(bold=True, size=12)
                ws_stats.cell(row=1, column=2).font = Font(bold=True, size=12)
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel report saved to: {output_path}")


def main():
    print(f"\n{'='*60}")
    print(f"STEP 4: Generating Excel Report")
    print(f"{'='*60}")
    
    # Read grading results from Step 3
    results_path = get_path("grading_output/grading_results.json")
    
    if not results_path.exists():
        print("ERROR: Grading results not found!")
        print("Run Step 3 first: python scripts/03_grade_answers.py")
        return
    
    with open(results_path, "r", encoding="utf-8") as f:
        results = json.load(f)
    
    print(f"Loaded {len(results)} student results")
    
    # Generate Excel
    output_path = get_path(config["paths"]["output_excel"])
    create_excel_report(results, output_path)
    
    print(f"\n{'='*60}")
    print("PIPELINE COMPLETE! 🎉")
    print(f"{'='*60}")
    print(f"\nOutput files:")
    print(f"  1. Cleaned files: cleaned_submissions/")
    print(f"  2. Renamed files: renamed_submissions/")
    print(f"  3. Grading data: grading_output/grading_results.json")
    print(f"  4. Excel report: {output_path}")
    print()


if __name__ == "__main__":
    main()